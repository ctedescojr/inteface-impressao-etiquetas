import tkinter as tk
from tkinter import ttk
from tkinter import font as tkfont
import os
import openpyxl
import pandas as pd
import datetime
import winsound
from datetime import datetime

# # Define o endereço IP e porta da impressora
# IP_IMPRESSORA = '169.254.70.214'
# PORTA_IMPRESSORA = 9100

# # Cria um socket TCP
# sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
# sock.connect((IP_IMPRESSORA, PORTA_IMPRESSORA))

# Define o caminho do arquivo onde as etiquetas serão salvas
file_path = os.path.abspath("./Base de Dados/Etiquetas.xlsx")

# Variável global que fará o controle dos baldes a serem impressos
global balde_atual

# Cores a serem usadas para o programaS
color1 = "#FF0000"      # vermelho
color1b = "#FF7171"     # vermeçho claro
color2 = "#007033"      # verde escuro
color2b = "#00B451"     # verde claro
color3 = "#FFFFFF"      # branco
color4 = "#0338E3"      # azul escuro
color4b = "#4671FC"     # azul claro
color5 = "#BC8F00"      # laranja escuro
color5b = "#FFD347"     # laranja claro

def createFile():
    cabecalho = ['Lote', 'Fornecedor', 'Florada', 'Balde', 'Cliente', 'NF', 'Data']
    new_file = openpyxl.Workbook()
    planilha = new_file.active
    planilha.append(cabecalho)
    new_file.save(file_path)

# Cria o arquivo para a base de dados no Excel caso ela ainda não exista
if not os.path.exists(file_path):
    createFile()

# Classe que será o container de todas as telas do programa
class mainFrame(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        self.title_font = tkfont.Font(family='Verdana', size=12, weight="bold", slant="roman")

        # o container será onde todas as telas serão alocadas
        # uma acima da outra, sendo que a escolhida para ser exibida
        # ficará acima das outras
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        self.geometry("804x604+350+100")
        self.title("Vida Natural - Impressão de Etiquetas")
        self.iconbitmap(default="./Templates/abelha.ico")
        self.wm_resizable(width=False, height=False)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for F in (WelcomePage, PrintPage, RePrintPage):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # Colocando todas as páginas no mesmo local;
            # a que ficar no topo, será a qual ficará visível.
            frame.grid(row=0, column=0, sticky="nsew")

        # Mostrando a tela de abertura
        self.show_frame("WelcomePage")

    # Função que fará a subida das telas escolhidas
    def show_frame(self, page_name):
        '''Mostra a página do nome identificado'''
        frame = self.frames[page_name]
        frame.tkraise()

    # Função para sair do programa e tocar um som de até logo 
    def exitProgram(self):
        winsound.PlaySound("./Templates/sair.wav", winsound.SND_FILENAME)
        self.quit()
# Classe da página de abertura do programa
class WelcomePage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.wpBackground = tk.PhotoImage(file="./Templates/mainBackground.png")
        label = tk.Label(self, image= self.wpBackground)
        label.place(x=0, y=0)

        # Função para o botão de ajuda da tela inicial do programa 
        def helpAbout():
            winsound.PlaySound("./Templates/ajuda.wav", winsound.SND_FILENAME)
            hA = tk.Toplevel(self)
            hA.title("Ajuda")
            # Posicionando a nova janela no centro da janela principal
            x_main = self.winfo_rootx()
            y_main = self.winfo_rooty()
            x_new = x_main + (804 - 504)//2
            y_new = y_main + (604 - 184)//2
            hA.geometry("504x184+{}+{}".format(x_new, y_new))
            #------#
            hA.iconbitmap("./Templates/ajuda.ico")
            help_label = tk.Label(hA, text="Instruções\n"
                                  "NOVO LOTE - Clique para imprimir um novo lote de etiquetas.\n"
                                  "LOTE ANTIGO - Clique para imprimir um lote antigo ou uma etiqueta antiga.\n"
                                  "CONSULTAR LOTES - Clique para conferir lotes de etiquetas já inseridos no sistema.\n"
                                  "SAIR - Clique para sair do programa.",
                                  font=('Arial 10'))
            
            about_label = tk.Label(hA, text="\n\nAutor: Clayton Del Tedesco Júnior\n"
                                   "Versão 1.0", font=('Arial 8'))
            help_label.pack()
            about_label.pack()

            button_ok = tk.Button(hA, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=hA.destroy)
            button_ok.pack(side='bottom')
            hA.grab_set()
            hA.focus()
            

        # Função que cria a base de etiquetas se ela ainda não existe, ou abre o arquivo para visualização no Excel
        def seeDataBase():
            # Cria janela avisando que o arquivo foi criado e o local onde ele está salvo
            arquivoCriadoBox = tk.Toplevel(self)
            arquivoCriadoBox.title("Aviso")
            # Posicionando a nova janela no centro da janela principal
            x_main = self.winfo_rootx()
            y_main = self.winfo_rooty()
            x_new = x_main + (804 - 704)//2
            y_new = y_main + (604 - 84)//2
            arquivoCriadoBox.geometry("704x84+{}+{}".format(x_new, y_new))

            button_ok = tk.Button(arquivoCriadoBox, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=arquivoCriadoBox.destroy)
            button_ok.pack(side='bottom')

            if os.path.exists(file_path):
                os.startfile(file_path)
                aviso_label = tk.Label(arquivoCriadoBox, text=f"Abrindo arquivo\n{file_path}", font=('Arial 10'))
            else:
                createFile()
                aviso_label = tk.Label(arquivoCriadoBox, text=f"{file_path}\nFoi criado com sucesso!", font=('Arial 10'))

            aviso_label.pack()
            arquivoCriadoBox.grab_set()
            arquivoCriadoBox.focus()

        # Botão para novo lote de etiquetas
        button_newBatch = tk.Button(self, text="Novo\nLote", width=8, height=2, bg=color2, fg=color3, activebackground=color2b, activeforeground=color3, 
                                    borderwidth=3, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: controller.show_frame("PrintPage"))
        # Botão para reimprimir etiquetas
        button_rePrint = tk.Button(self, text="Lote\nAntigo", width=8, height=2, bg=color2, fg=color3, activebackground=color2b, activeforeground=color3,
                                   borderwidth=3, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: controller.show_frame("RePrintPage"))
        # Botão para consultar etiquetas gravadas
        button_seeDataBase = tk.Button(self, text="Consultar\nLotes", width=9, height=4, bg=color2, fg=color3, activebackground=color2b, activeforeground=color3,
                                       borderwidth=3, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: seeDataBase())
        # Botão para ajuda e informações sobre o programa
        button_helpAbout = tk.Button(self, text="?", width=2, height=1, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3,
                                     borderwidth=2, font=('Ive 16 bold'), relief='raised', overrelief='ridge', command=lambda: helpAbout())
        # Botão para sair
        button_exit = tk.Button(self, text="Sair", width=7, height=1, bg=color1, fg=color3, activebackground=color1b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: mainFrame.exitProgram(self))

        # Posicionamento dos botões na tela
        button_newBatch.place(x=246, y=379)
        button_rePrint.place(x=370, y=379)
        button_seeDataBase.place(x=506, y=335)
        button_helpAbout.place(x=750, y=20)
        button_exit.place(x=690, y=542)

class PrintPage(tk.Frame):
    global balde_atual
    balde_atual = '🍯'     # Seta a o balde para um pomte de mel, indicando que não existe nada a ser impresso

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.wpBackground = tk.PhotoImage(file="./Templates/printPage.png")
        labelb = tk.Label(self, image= self.wpBackground)
        labelb.place(x=0, y=0)

        # Função para o botão de ajuda da tela de criação de novas etiquetas a serem impressas
        def helpPrint():
            winsound.PlaySound("./Templates/ajuda.wav", winsound.SND_FILENAME)
            hP = tk.Toplevel(self)
            hP.title("Ajuda")
            # Posicionando a nova janela no centro da janela principal
            x_main = self.winfo_rootx()
            y_main = self.winfo_rooty()
            x_new = x_main + (804 - 704)//2
            y_new = y_main + (604 - 184)//2
            hP.geometry("704x184+{}+{}".format(x_new, y_new))
            #-----------#
            hP.iconbitmap("./Templates/ajuda.ico")
            help_label = tk.Label(hP, text="Instruções\n"
                                  "Preencha os campos de fornecedor, florada, n° baldes, cliente, lote e nota fiscal.\n"
                                  "CRIAR ETIQUETAS - Clique para criar as etiquetas com os campos preenchidos e salvar as etiquetas no sistema.\n"
                                  "IMPRIMIR - Clique para imprimir a etiqueta mostrada no campo Etiqueta Atual.\n"
                                  "REIMPRIMIR - Digite o balde que deseja reimprimir e clique em Reimprimir.\n"
                                  "VOLTAR - Clique para voltar a tela inicial do programa.\n"
                                  "SAIR - Clique para sair do programa.",
                                  font=('Arial 10'))
            
            help_label.pack()

            button_ok = tk.Button(hP, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=hP.destroy)
            button_ok.pack(side='bottom')
            hP.grab_set()
            hP.focus()

        # ---------- Criação dos campos de entrada -------- #
        input_lote = ttk.Entry(self, font=('Arial 15'), width=8)
        input_fornecedor = ttk.Entry(self, font=('Arial 15'), width=26)
        input_florada = ttk.Entry(self, font=('Arial 15'), width=20)
        input_nBaldes = ttk.Entry(self, font=('Arial 15'), width=4, justify='center')
        input_cliente = ttk.Entry(self, font=('Arial 15'), width=30)
        input_NF = ttk.Entry(self, font=('Arial 15'), width=10)
        input_reimprimirBalde = ttk.Entry(self, font=('Arial 15'), width=4, justify='center')
        #---------------------------------------------------#
        # ---------- Posicionamento dos campos de entrada ---------- #
        input_fornecedor.place(x=330, y=158)
        input_florada.place(x=363, y=207)
        input_nBaldes.place(x=450, y=257)
        input_cliente.place(x=308, y=308)
        input_lote.place(x=430, y=355)
        input_NF.place(x=419, y=405)
        input_reimprimirBalde.place(x=702, y=233)
        #------------------------------------------------------------#
        lista_etiquetas = []

        def criarEtiquetas():
            global balde_atual

            baseEtiquetas = openpyxl.load_workbook(file_path)
            planilha = baseEtiquetas.active
            sheet = baseEtiquetas['Sheet']
            # planilha.column_dimensions['A'].width = 6
            # planilha['A1'].number_format = '@'

            fornecedor = input_fornecedor.get()
            florada = input_florada.get()
            nBaldes = input_nBaldes.get()
            cliente = input_cliente.get()
            lote = input_lote.get()
            nf = input_NF.get()

            # Verifica se o lote é válido pela quantidade de caracteres: 6
            if len(lote) != 6:
                # Cria uma janela avisando que a quantidade de carecteres do lote está errado
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroLote = tk.Toplevel(self)
                erroLote.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroLote.geometry("404x84+{}+{}".format(x_new, y_new))
                erroLote.iconbitmap("./Templates/erro.ico")
                
                erroLote_label = tk.Label(erroLote, text="Não foi possível criar as etiquetas.\nVerifique se digitou os 6 dígitos do lote.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroLote, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroLote.destroy)
                erroLote_label.pack()
                button_ok.pack(side='bottom')
                erroLote.grab_set()
                erroLote.focus()
                return

            # Verifica se o lote já foi inserido anteriormente
            for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
                if row[0].value == lote:
                    # Cria janela de erro caso o lote já esteja presente no sistema
                    winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                    erroBox1 = tk.Toplevel(self)
                    erroBox1.title("Erro")
                    # Posicionando a nova janela no centro da janela principal
                    x_main = self.winfo_rootx()
                    y_main = self.winfo_rooty()
                    x_new = x_main + (804 - 404)//2
                    y_new = y_main + (604 - 84)//2
                    erroBox1.geometry("404x84+{}+{}".format(x_new, y_new))
                    erroBox1.iconbitmap("./Templates/erro.ico")
                
                    erro1_label = tk.Label(erroBox1, text="Não foi possível criar as etiquetas.\nLote já existente no sistema.\n",
                                        font = ('Arial 10'))
                    button_ok = tk.Button(erroBox1, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                        borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox1.destroy)
                    erro1_label.pack()
                    button_ok.pack(side='bottom')
                    erroBox1.grab_set()
                    erroBox1.focus()
                    return
            
            if(len(fornecedor) == 0 or len(florada) == 0 or len(nBaldes) == 0 or len(cliente) == 0 or len(lote) == 0):
                # Cria uma janela avisando que possui campos vazios
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox2 = tk.Toplevel(self)
                erroBox2.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroBox2.geometry("404x84+{}+{}".format(x_new, y_new))
                erroBox2.iconbitmap("./Templates/erro.ico")
                
                erro2_label = tk.Label(erroBox2, text="Não foi possível criar as etiquetas.\nVerifique se todos os campos foram preenchidos\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox2, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox2.destroy)
                erro2_label.pack()
                button_ok.pack(side='bottom')
                erroBox2.grab_set()
                erroBox2.focus()

            elif not nBaldes.isnumeric():
                # Cria uma janela avisando que foi digitado errado a quantidade de baldes
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox3 = tk.Toplevel(self)
                erroBox3.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroBox3.geometry("404x84+{}+{}".format(x_new, y_new))
                erroBox3.iconbitmap("./Templates/erro.ico")
                
                erro3_label = tk.Label(erroBox3, text="Não foi possível criar as etiquetas.\nDigite um número válido para a quantidade de baldes.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox3, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox3.destroy)
                erro3_label.pack()
                button_ok.pack(side='bottom')
                erroBox3.grab_set()
                erroBox3.focus()
                
            else:
                for contBalde in range(int(nBaldes)):
                    etiquetas = {
                        "Lote": lote,
                        "Fornecedor": fornecedor,
                        "Florada": florada,
                        "Balde": contBalde + 1,
                        "Cliente": cliente,
                        "NF": nf
                    }
                    data = datetime.now()
                    data_formatada = data.strftime('%d-%m-%Y')
                    lista_etiquetas.append(etiquetas)
                    to_append = [lote, fornecedor, florada, contBalde + 1,  cliente, nf, data_formatada]
                    planilha.append(to_append)
                baseEtiquetas.save(file_path)


                with open("Mensagem.msg", "w", encoding='ansi', newline='\r\n') as arquivo:
                    arquivo.write(f"Fornecedor: {etiquetas['Fornecedor']}\n")
                    arquivo.write(f"Florada: {etiquetas['Florada']}\n")
                    arquivo.write(f"Cliente: {etiquetas['Cliente']}\n")
                    arquivo.write(f"N. Baldes:      / {nBaldes}\n")
                    arquivo.write(f'Data Recebimento:{data_formatada}\n')
                    arquivo.write(f"Lote: {etiquetas['Lote']}\n")
                    arquivo.write(f"NF: {etiquetas['NF']}\n\n")

            
                balde_atual = 1                     # Seta o balde atual para 1, indicando que já pode ser impresso
                label_balde['text'] = balde_atual   # Mostra qual balde está para ser impresso
                # Cria uma janela avisando que as etiquetas foram criadas
                criadasBox = tk.Toplevel(self)
                criadasBox.title("Sucesso!")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                criadasBox.geometry("404x84+{}+{}".format(x_new, y_new))
                criadas_label = tk.Label(criadasBox, text=f"Foram criadas {nBaldes} etiquetas!\n",
                                         font = ('Arial 10'))
                button_ok = tk.Button(criadasBox, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                      borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command= criadasBox.destroy)
                criadas_label.pack()
                button_ok.pack(side='bottom')
                criadasBox.grab_set()
                criadasBox.focus()

        def imprimeEtiqueta():
            global balde_atual
            # Verifica se já foi realizada a criação das etiquetas
            if len(lista_etiquetas) == 0:
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox1 = tk.Toplevel(self)
                erroBox1.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroBox1.geometry("404x84+{}+{}".format(x_new, y_new))
                erroBox1.iconbitmap("./Templates/erro.ico")
                
                erroBox1_label = tk.Label(erroBox1, text="Não foi possível imprimir as etiquetas.\nCertifique-se que foram criadas com o botão Criar Etiquetas\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox1, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox1.destroy)
                erroBox1_label.pack()
                button_ok.pack(side='bottom')
                erroBox1.grab_set()
                erroBox1.focus()
                return

            if balde_atual > len(lista_etiquetas):
                # Cria janela avisando que já foram impressas todas as etiquetas
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox2 = tk.Toplevel(self)
                erroBox2.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroBox2.geometry("404x84+{}+{}".format(x_new, y_new))
                erroBox2.iconbitmap("./Templates/erro.ico")
                
                erro2_label = tk.Label(erroBox2, text="Não foi possível imprimir uma nova etiqueta.\nTodas as etiquetas já foram impressas.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox2, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox2.destroy)
                erro2_label.pack()
                label_balde['text'] = len(lista_etiquetas)
                button_ok.pack(side='bottom')
                erroBox2.grab_set()
                erroBox2.focus()

            else:
                print(lista_etiquetas[balde_atual-1])
                balde_atual += 1                    # Atualiza o balde a ser impresso
                if(balde_atual > len(lista_etiquetas)):
                    
                    label_balde['fg'] = color2
                    label_balde['text'] = len(lista_etiquetas)   # Mostra o último balde impresso
                    # Cria uma janela avisando que as etiquetas foram criadas
                    impressasBox = tk.Toplevel(self)
                    impressasBox.title("Sucesso!")
                    # Posicionando a nova janela no centro da janela principal
                    x_main = self.winfo_rootx()
                    y_main = self.winfo_rooty()
                    x_new = x_main + (804 - 404)//2
                    y_new = y_main + (604 - 84)//2
                    impressasBox.geometry("404x84+{}+{}".format(x_new, y_new))
                    impressas_label = tk.Label(impressasBox, text=f"Foram impressas {len(lista_etiquetas)} etiquetas!\n",
                                            font = ('Arial 10'))
                    button_ok = tk.Button(impressasBox, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                        borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=impressasBox.destroy)
                    impressas_label.pack()
                    button_ok.pack(side='bottom')
                    impressasBox.grab_set()
                    impressasBox.focus()
                    winsound.PlaySound("./Templates/success.wav", winsound.SND_FILENAME)
                else:
                    label_balde['text'] = balde_atual   # Mostra qual balde está para ser impresso

        # Função que reimprime um balde presente no lote
        def reimprimeEtiqueta():
            # Verifica se existe uma lista de etiquetas a ser impressa
            if len(lista_etiquetas) == 0:
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox1 = tk.Toplevel(self)
                erroBox1.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroBox1.geometry("404x84+{}+{}".format(x_new, y_new))
                erroBox1.iconbitmap("./Templates/erro.ico")
                
                erroBox1_label = tk.Label(erroBox1, text="Não foi possível reimprimir a etiqueta.\nNenhuma etiqueta foi impressa ainda.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox1, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox1.destroy)
                erroBox1_label.pack()
                button_ok.pack(side='bottom')
                erroBox1.grab_set()
                erroBox1.focus()
                return

            reimprimeBalde = input_reimprimirBalde.get()

            if not reimprimeBalde.isnumeric():
                # Cria uma janela avisando que foi digitado errado a quantidade de baldes
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox3 = tk.Toplevel(self)
                erroBox3.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroBox3.geometry("404x84+{}+{}".format(x_new, y_new))
                erroBox3.iconbitmap("./Templates/erro.ico")
                
                erro3_label = tk.Label(erroBox3, text="Não foi possível reimprimir a etiqueta."
                                       f"\nDigite um número válido entre 1 e {len(lista_etiquetas)} para a quantidade de baldes.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox3, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox3.destroy)
                erro3_label.pack()
                button_ok.pack(side='bottom')
                erroBox3.grab_set()
                erroBox3.focus()
                return
                
            reimprimeBalde = int(reimprimeBalde)
            if reimprimeBalde > len(lista_etiquetas):
                # Cria uma janela avisando que foi digitado errado a quantidade de baldes
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox3 = tk.Toplevel(self)
                erroBox3.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 504)//2
                y_new = y_main + (604 - 84)//2
                erroBox3.geometry("504x84+{}+{}".format(x_new, y_new))
                erroBox3.iconbitmap("./Templates/erro.ico")
                
                erro3_label = tk.Label(erroBox3, text="Não foi possível reimprimir a etiqueta. Número inválido para a quantidade de baldes."
                                       f"\nDigite um número válido entre 1 e {len(lista_etiquetas)} para a etiqueta a ser reimpressa.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox3, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox3.destroy)
                erro3_label.pack()
                button_ok.pack(side='bottom')
                erroBox3.grab_set()
                erroBox3.focus()
                
            else:
                print(lista_etiquetas[reimprimeBalde-1])

        # Função para limpar a tela de inserção de novo lote
        def limpar():
            input_fornecedor.delete(0, tk.END)
            input_florada.delete(0, tk.END)
            input_nBaldes.delete(0, tk.END)
            input_cliente.delete(0, tk.END)
            input_lote.delete(0, tk.END)
            input_NF.delete(0, tk.END)
            lista_etiquetas.clear()

        
        label_balde = tk.Label(self, text=balde_atual, font=('Ive 18 bold'), bg=color3, width=2, justify='center')

        # Botão para novo lote de etiquetas
        button_createStickers = tk.Button(self, text="Criar\nEtiquetas", width=8, height=2, bg=color2, fg=color3, activebackground=color2b, activeforeground=color3, 
                                          borderwidth=3, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: criarEtiquetas())
        
        button_print = tk.Button(self, text="Imprimir", width=10, height=2, bg=color2, fg=color3, activebackground=color2b, activeforeground=color3, 
                                 borderwidth=3, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: imprimeEtiqueta())
        
        button_rePrint = tk.Button(self, text="Reimprimir", width=10, height=1, bg=color2, fg=color3, activebackground=color2b, activeforeground=color3, 
                                 borderwidth=3, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: reimprimeEtiqueta())
        
        button_limpar = tk.Button(self, text="Limpar", width=7, height=1, bg=color1, fg=color3, activebackground=color1b, activeforeground=color3, 
                                 borderwidth=3, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: limpar())

        # Botão para ajuda e informações sobre o programa
        button_helpAbout = tk.Button(self, text="?", width=2, height=1, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3,
                                     borderwidth=2, font=('Ive 16 bold'), relief='raised', overrelief='ridge', command=lambda: helpPrint())

        # Botão para voltar
        button_back = tk.Button(self, text="Voltar", width=7, height=1, bg=color5, fg=color3, activebackground=color5b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: controller.show_frame("WelcomePage"))
        
        # Botão para sair
        button_exit = tk.Button(self, text="Sair", width=7, height=1, bg=color1, fg=color3, activebackground=color1b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: mainFrame.exitProgram(self))
        
        label_balde.place(x=586, y=488)
        button_createStickers.place(x=179, y=475)
        button_print.place(x = 412, y=475)
        button_rePrint.place(x = 662, y=283)
        button_limpar.place(x = 679, y = 345)
        button_helpAbout.place(x=750, y=20)
        button_back.place(x=23, y=542)
        button_exit.place(x=690, y=542)

class RePrintPage(tk.Frame):
    global balde_atual
    balde_atual = '🍯'     # Seta a o balde para um pomte de mel, indicando que não existe nada a ser impresso
    

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.wpBackground = tk.PhotoImage(file="./Templates/rePrintPage.png")
        labelb = tk.Label(self, image= self.wpBackground)
        labelb.place(x=0, y=0)

        # Função para o botão de ajuda da tela inicial do programa 
        def helpReprint():
            winsound.PlaySound("./Templates/ajuda.wav", winsound.SND_FILENAME)
            hA = tk.Toplevel(self)
            hA.title("Ajuda")
            # Posicionando a nova janela no centro da janela principal
            x_main = self.winfo_rootx()
            y_main = self.winfo_rooty()
            x_new = x_main + (804 - 604)//2
            y_new = y_main + (604 - 184)//2
            hA.geometry("604x184+{}+{}".format(x_new, y_new))
            #------#
            hA.iconbitmap("./Templates/ajuda.ico")
            help_label = tk.Label(hA, text="Instruções\n"
                                  "Digite o número do lote do qual deseja reimprimir as etiquetas e clique em Carregar.\n"
                                  "IMPRIMIR - Imprime o balde atual sequencialmente, use para reimprimir um lote inteiro\n"
                                  "REIMPRIMIR - Digite o balde que deseja reimprimir e clique para reimprimí-lo.\n"
                                  "VOLTAR - Clique para voltar à tela inicial.\n"
                                  "SAIR - Clique para sair do programa.",
                                  font=('Arial 10'))
            
            help_label.pack()

            button_ok = tk.Button(hA, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=hA.destroy)
            button_ok.pack(side='bottom')
            hA.grab_set()
            hA.focus()

        # ---------- Labels para carregamento das informações do lote ---------- #
        label_fornecedor = tk.Label(self, text=balde_atual, font=('Arial 14'), bg=color3, width=26)
        label_florada = tk.Label(self, text=balde_atual, font=('Arial 14'), bg=color3, width=20)
        label_nbalde = tk.Label(self, text=balde_atual, font=('Arial 14'), bg=color3, width=4, justify='left')
        label_cliente = tk.Label(self, text=balde_atual, font=('Arial 14'), bg=color3, width=28)
        label_lote = tk.Label(self, text=balde_atual, font=('Arial 14'), bg=color3, width=8)
        label_nf = tk.Label(self, text=balde_atual, font=('Arial 14'), bg=color3, width=10)
        label_data = tk.Label(self, text=balde_atual, font=('Arial 14'), bg=color3, width=10)
        label_peso = tk.Label(self, text=balde_atual, font=('Arial 14'), bg=color3, width=5)
        label_balde = tk.Label(self, text=balde_atual, font=('Ive 18 bold'), bg=color3, width=2, justify='center')
        #------------------------------------------------------------------------#

        # ---------- Posicionamento das labes carregadas ---------- #
        label_balde.place(x=586, y=488)
        label_fornecedor.place(x=325, y=180)
        label_florada.place(x=325, y=218)
        label_nbalde.place(x=325, y=255)
        label_cliente.place(x=325, y=292)
        label_lote.place(x=325, y=330)
        label_nf.place(x=325, y=367)
        label_data.place(x=325, y=404)
        label_peso.place(x=540, y=404)
        #-----------------------------------------------------------#

        # ---------- Criação dos campos de entrada -------- #
        input_lote = ttk.Entry(self, font=('Arial 15'), width=8)
        input_reimprimirBalde = ttk.Entry(self, font=('Arial 15'), width=4, justify='center')
       
        # ---------- Posicionamento dos campos de entrada ---------- #
        input_lote.place(x=333, y=130)
        input_reimprimirBalde.place(x=702, y=233)

        lista_etiquetas = []    # Lista para carregar o lote solicitado

        def carregarLote():

            if lista_etiquetas:
                lista_etiquetas.clear()
                
            global balde_atual
            df_lote = pd.read_excel(file_path, dtype={'Lote':str})

            numLote = input_lote.get()

            # Verifica se o lote é válido pela quantidade de caracteres: 6
            if len(numLote) != 6:
                # Cria uma janela avisando que a quantidade de carecteres do lote está errado
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroLote = tk.Toplevel(self)
                erroLote.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroLote.geometry("404x84+{}+{}".format(x_new, y_new))
                erroLote.iconbitmap("./Templates/erro.ico")
                
                erroLote_label = tk.Label(erroLote, text="Não foi possível carregar as etiquetas.\nVerifique se digitou os 6 dígitos do lote.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroLote, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroLote.destroy)
                erroLote_label.pack()
                button_ok.pack(side='bottom')
                erroLote.grab_set()
                erroLote.focus()
                return
            
            # Verifica se o lote existe no arquivo carregado
            if str(numLote) not in df_lote['Lote'].unique():
                # Cria uma janela avisando que o lote não existe
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroLote = tk.Toplevel(self)
                erroLote.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroLote.geometry("404x84+{}+{}".format(x_new, y_new))
                erroLote.iconbitmap("./Templates/erro.ico")
                
                erroLote_label = tk.Label(erroLote, text="Não foi possível carregar as etiquetas.\nLote não existe.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroLote, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroLote.destroy)
                erroLote_label.pack()
                button_ok.pack(side='bottom')
                erroLote.grab_set()
                erroLote.focus()
                return
            

            else:
                df_lote_filtrado = df_lote[df_lote['Lote'] == numLote]

                lista_dicionario = df_lote_filtrado.to_dict('records')

                for dicionario_etiqueta in lista_dicionario:
                    info_etiqueta = (dicionario_etiqueta['Lote'], dicionario_etiqueta['Fornecedor'], dicionario_etiqueta['Florada'], dicionario_etiqueta['Balde'],
                                    dicionario_etiqueta['Cliente'], dicionario_etiqueta['NF'], dicionario_etiqueta['Data'])
                    lista_etiquetas.append(info_etiqueta)

                balde_atual = 1                     # Seta o balde atual para 1, indicando que já pode ser impresso
                label_balde['text'] = balde_atual   # Mostra qual balde está para ser impresso

                label_lote['text'] = lista_etiquetas[0][0]
                label_fornecedor['text'] = lista_etiquetas[0][1]
                label_florada['text'] = lista_etiquetas[0][2]
                label_nbalde['text'] = len(lista_etiquetas)
                label_cliente['text'] = lista_etiquetas[0][4]
                label_nf['text'] = lista_etiquetas[0][5]
                label_data['text'] = lista_etiquetas[0][6]
                
                with open("Mensagem.msg", "w", encoding='ansi', newline='\r\n') as arquivo:
                    arquivo.write(f"Fornecedor: {lista_etiquetas[0][1]}\n")
                    arquivo.write(f"Florada: {lista_etiquetas[0][2]}\n")
                    arquivo.write(f"Cliente: {lista_etiquetas[0][4]}\n")
                    arquivo.write(f"N. Baldes:      / {len(lista_etiquetas)}\n")
                    arquivo.write(f'Data Recebimento:{lista_etiquetas[0][6]}\n')
                    arquivo.write(f"Lote: {lista_etiquetas[0][0]}\n")
                    arquivo.write(f"NF: {lista_etiquetas[0][5]}\n\n")

        def imprimeEtiqueta():
            global balde_atual
            # Verifica se já foi realizada a criação das etiquetas
            if len(lista_etiquetas) == 0:
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox1 = tk.Toplevel(self)
                erroBox1.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroBox1.geometry("404x84+{}+{}".format(x_new, y_new))
                erroBox1.iconbitmap("./Templates/erro.ico")
                
                erroBox1_label = tk.Label(erroBox1, text="Não foi possível imprimir as etiquetas.\nCertifique-se que o lote foi carregado.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox1, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox1.destroy)
                erroBox1_label.pack()
                button_ok.pack(side='bottom')
                erroBox1.grab_set()
                erroBox1.focus()
                return

            if balde_atual > len(lista_etiquetas):
                # Cria janela avisando que já foram impressas todas as etiquetas
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox2 = tk.Toplevel(self)
                erroBox2.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroBox2.geometry("404x84+{}+{}".format(x_new, y_new))
                erroBox2.iconbitmap("./Templates/erro.ico")
                
                erro2_label = tk.Label(erroBox2, text="Não foi possível imprimir uma nova etiqueta.\nTodas as etiquetas já foram impressas.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox2, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox2.destroy)
                erro2_label.pack()
                label_balde['text'] = len(lista_etiquetas)
                button_ok.pack(side='bottom')
                erroBox2.grab_set()
                erroBox2.focus()

            else:
                print(lista_etiquetas[balde_atual-1])
                balde_atual += 1                    # Atualiza o balde a ser impresso
                if(balde_atual > len(lista_etiquetas)):
                    
                    label_balde['fg'] = color2
                    label_balde['text'] = len(lista_etiquetas)   # Mostra o último balde impresso
                    # Cria uma janela avisando que as etiquetas foram criadas
                    impressasBox = tk.Toplevel(self)
                    impressasBox.title("Sucesso!")
                    # Posicionando a nova janela no centro da janela principal
                    x_main = self.winfo_rootx()
                    y_main = self.winfo_rooty()
                    x_new = x_main + (804 - 404)//2
                    y_new = y_main + (604 - 84)//2
                    impressasBox.geometry("404x84+{}+{}".format(x_new, y_new))
                    impressas_label = tk.Label(impressasBox, text=f"Foram impressas {len(lista_etiquetas)} etiquetas!\n",
                                            font = ('Arial 10'))
                    button_ok = tk.Button(impressasBox, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                        borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=impressasBox.destroy)
                    impressas_label.pack()
                    button_ok.pack(side='bottom')
                    impressasBox.grab_set()
                    impressasBox.focus()
                    winsound.PlaySound("./Templates/success.wav", winsound.SND_FILENAME)
                else:
                    label_balde['text'] = balde_atual   # Mostra qual balde está para ser impresso

        def reimprimeEtiqueta():
            # Verifica se existe uma lista de etiquetas a ser impressa
            if len(lista_etiquetas) == 0:
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox1 = tk.Toplevel(self)
                erroBox1.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroBox1.geometry("404x84+{}+{}".format(x_new, y_new))
                erroBox1.iconbitmap("./Templates/erro.ico")
                
                erroBox1_label = tk.Label(erroBox1, text="Não foi possível reimprimir a etiqueta.\nNenhuma etiqueta foi impressa ainda.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox1, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox1.destroy)
                erroBox1_label.pack()
                button_ok.pack(side='bottom')
                erroBox1.grab_set()
                erroBox1.focus()
                return

            reimprimeBalde = input_reimprimirBalde.get()

            if not reimprimeBalde.isnumeric():
                # Cria uma janela avisando que foi digitado errado a quantidade de baldes
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox3 = tk.Toplevel(self)
                erroBox3.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 404)//2
                y_new = y_main + (604 - 84)//2
                erroBox3.geometry("404x84+{}+{}".format(x_new, y_new))
                erroBox3.iconbitmap("./Templates/erro.ico")
                
                erro3_label = tk.Label(erroBox3, text="Não foi possível reimprimir a etiqueta."
                                       f"\nDigite um número válido entre 1 e {len(lista_etiquetas)} para a quantidade de baldes.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox3, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox3.destroy)
                erro3_label.pack()
                button_ok.pack(side='bottom')
                erroBox3.grab_set()
                erroBox3.focus()
                return
                
            reimprimeBalde = int(reimprimeBalde)
            if reimprimeBalde > len(lista_etiquetas):
                # Cria uma janela avisando que foi digitado errado a quantidade de baldes
                winsound.PlaySound("./Templates/erro.wav", winsound.SND_FILENAME)
                erroBox3 = tk.Toplevel(self)
                erroBox3.title("Erro")
                # Posicionando a nova janela no centro da janela principal
                x_main = self.winfo_rootx()
                y_main = self.winfo_rooty()
                x_new = x_main + (804 - 504)//2
                y_new = y_main + (604 - 84)//2
                erroBox3.geometry("504x84+{}+{}".format(x_new, y_new))
                erroBox3.iconbitmap("./Templates/erro.ico")
                
                erro3_label = tk.Label(erroBox3, text="Não foi possível reimprimir a etiqueta. Número inválido para a quantidade de baldes."
                                       f"\nDigite um número válido entre 1 e {len(lista_etiquetas)} para a etiqueta a ser reimpressa.\n",
                                      font = ('Arial 10'))
                button_ok = tk.Button(erroBox3, text="OK", width=7, height=0, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 12 bold'), relief='raised', overrelief='ridge', command=erroBox3.destroy)
                erro3_label.pack()
                button_ok.pack(side='bottom')
                erroBox3.grab_set()
                erroBox3.focus()
                
            else:
                print(lista_etiquetas[reimprimeBalde-1])


        # Botão para ajuda e informações sobre o programa
        button_helpAbout = tk.Button(self, text="?", width=2, height=1, bg=color4, fg=color3, activebackground=color4b, activeforeground=color3,
                                     borderwidth=2, font=('Ive 16 bold'), relief='raised', overrelief='ridge', command=lambda: helpReprint())
        
        # Botão para carregar o lote desejado
        button_carregar = tk.Button(self, text="Carregar", width=10, height=1, bg=color2, fg=color3, activebackground=color2b, activeforeground=color3, 
                                 borderwidth=3, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: carregarLote())

        # Botão para imprimir as etiquetas carregas
        button_print = tk.Button(self, text="Imprimir", width=10, height=2, bg=color2, fg=color3, activebackground=color2b, activeforeground=color3, 
                                 borderwidth=3, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: imprimeEtiqueta())
        
        # Botão para reimprimir alguma etiqueta específica
        button_rePrint = tk.Button(self, text="Reimprimir", width=10, height=1, bg=color2, fg=color3, activebackground=color2b, activeforeground=color3, 
                                 borderwidth=3, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: reimprimeEtiqueta())
        
        # Botão para voltar
        button_back = tk.Button(self, text="Voltar", width=7, height=1, bg=color5, fg=color3, activebackground=color5b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: controller.show_frame("WelcomePage"))

        # Botão para sair
        button_exit = tk.Button(self, text="Sair", width=7, height=1, bg=color1, fg=color3, activebackground=color1b, activeforeground=color3, 
                                borderwidth=2, font=('Ive 14 bold'), relief='raised', overrelief='ridge', command=lambda: mainFrame.exitProgram(self))
        
        button_helpAbout.place(x=750, y=20)
        button_carregar.place(x=440, y=125)
        button_print.place(x = 412, y=475)
        button_rePrint.place(x = 662, y=283)
        button_back.place(x=25, y=542)
        button_exit.place(x=690, y=542)

if __name__ == "__main__":
    app = mainFrame()
    app.mainloop()
